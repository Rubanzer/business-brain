"""Data Engineer Agent — parse, clean, load, and auto-generate metadata for uploaded files."""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from datetime import datetime
from typing import Any

from google import genai
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from business_brain.ingestion.context_ingestor import ingest_context
from business_brain.memory import metadata_store
from config.settings import settings

logger = logging.getLogger(__name__)

_client: genai.Client | None = None

_DATE_PATTERNS = [
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "%Y-%m-%d"),
    (re.compile(r"^\d{2}/\d{2}/\d{4}$"), "%m/%d/%Y"),
    (re.compile(r"^\d{2}-\d{2}-\d{4}$"), "%m-%d-%Y"),
    (re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}"), "%Y-%m-%dT%H:%M"),
]


def _get_client() -> genai.Client:
    global _client
    if _client is None:
        _client = genai.Client(api_key=settings.gemini_api_key)
    return _client


# ---------------------------------------------------------------------------
# File parsing (lightweight — no pandas/numpy)
# ---------------------------------------------------------------------------

def parse_csv(raw: bytes) -> list[dict]:
    """Parse CSV bytes into a list of row dicts using stdlib csv."""
    text_io = io.StringIO(raw.decode("utf-8-sig"))
    reader = csv.DictReader(text_io)
    return [dict(row) for row in reader]


def parse_excel(raw: bytes) -> list[dict]:
    """Parse Excel bytes into a list of row dicts using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    result = []
    for row in rows_iter:
        result.append({h: (cell if cell is not None else "") for h, cell in zip(headers, row)})
    wb.close()
    return result


def parse_pdf(raw: bytes) -> str:
    """Extract all text from a PDF. Returns plain text (not tabular rows)."""
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    parts = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            parts.append(t)
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Type inference (no numpy — pure Python)
# ---------------------------------------------------------------------------

_BOOL_TRUE = {"true", "yes", "t", "y"}
_BOOL_FALSE = {"false", "no", "f", "n"}


def _try_parse_type(value: str) -> str:
    """Return the most specific type tag for a single string value."""
    v = value.strip()
    if not v:
        return "empty"

    # Integer (check before boolean so "1"/"0" are treated as numbers)
    try:
        int(v)
        return "integer"
    except ValueError:
        pass

    # Float
    try:
        float(v)
        return "float"
    except ValueError:
        pass

    # Boolean (after numeric checks to avoid "1"/"0" ambiguity)
    if v.lower() in _BOOL_TRUE | _BOOL_FALSE:
        return "boolean"

    # Date
    for pattern, fmt in _DATE_PATTERNS:
        if pattern.match(v):
            try:
                datetime.strptime(v[:len(fmt)+2], fmt)
                return "date"
            except ValueError:
                pass

    return "text"


_TYPE_TO_PG = {
    "integer": "BIGINT",
    "float": "DOUBLE PRECISION",
    "boolean": "BOOLEAN",
    "date": "TIMESTAMP",
    "text": "TEXT",
    "empty": "TEXT",
}

# Priority: higher number wins when merging column types
_TYPE_PRIORITY = {"empty": 0, "boolean": 1, "integer": 2, "float": 3, "date": 4, "text": 5}


def infer_column_types(rows: list[dict], sample_size: int = 100) -> dict[str, str]:
    """Infer PostgreSQL column types from sample rows.

    Returns mapping of column_name → PG type string.
    """
    if not rows:
        return {}

    sample = rows[:sample_size]
    col_types: dict[str, str] = {}

    for col in rows[0]:
        detected: dict[str, int] = {}
        for row in sample:
            val = str(row.get(col, ""))
            t = _try_parse_type(val)
            detected[t] = detected.get(t, 0) + 1

        # Pick the type with highest priority among those seen (ignoring empty)
        non_empty = {t: c for t, c in detected.items() if t != "empty"}
        if not non_empty:
            col_types[col] = "TEXT"
        else:
            winner = max(non_empty, key=lambda t: _TYPE_PRIORITY.get(t, 5))
            col_types[col] = _TYPE_TO_PG.get(winner, "TEXT")

    return col_types


# ---------------------------------------------------------------------------
# Data hygiene checks
# ---------------------------------------------------------------------------

def check_hygiene(rows: list[dict]) -> list[dict]:
    """Analyze data hygiene issues. Returns list of issue dicts."""
    if not rows:
        return []

    issues: list[dict] = []
    columns = list(rows[0].keys())
    total = len(rows)

    for col in columns:
        values = [str(row.get(col, "")) for row in rows]

        # Null / empty count
        empty_count = sum(1 for v in values if v.strip() == "")
        completeness = round((total - empty_count) / total * 100, 1)
        if empty_count > 0:
            issues.append({
                "column": col,
                "issue": f"{empty_count}/{total} empty values ({completeness}% complete)",
                "action": "will_coerce_empty_to_null",
            })

        # Whitespace issues
        ws_count = sum(1 for v in values if v != v.strip() and v.strip() != "")
        if ws_count > 0:
            issues.append({
                "column": col,
                "issue": f"{ws_count} values with leading/trailing whitespace",
                "action": "will_strip_whitespace",
            })

        # Type mismatch detection
        non_empty = [v.strip() for v in values if v.strip() != ""]
        if non_empty:
            types_seen = {_try_parse_type(v) for v in non_empty[:100]}
            types_seen.discard("empty")
            if len(types_seen) > 1 and "text" not in types_seen:
                issues.append({
                    "column": col,
                    "issue": f"mixed types detected: {', '.join(sorted(types_seen))}",
                    "action": "will_coerce_mismatches_to_null",
                })

    # Duplicate rows
    seen: set[tuple] = set()
    dup_count = 0
    for row in rows:
        key = tuple(str(row.get(c, "")) for c in columns)
        if key in seen:
            dup_count += 1
        seen.add(key)

    if dup_count > 0:
        issues.append({
            "column": "*",
            "issue": f"{dup_count} exact duplicate rows",
            "action": "will_remove_duplicates",
        })

    return issues


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------

def _drop_empty_columns(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Remove columns that are entirely empty across all rows.

    Returns (rows_with_columns_removed, list_of_dropped_column_names).
    """
    if not rows:
        return rows, []

    columns = list(rows[0].keys())
    empty_cols: list[str] = []
    for col in columns:
        if all(str(row.get(col, "")).strip() == "" or row.get(col) is None for row in rows):
            empty_cols.append(col)

    if not empty_cols:
        return rows, []

    drop_set = set(empty_cols)
    cleaned = [{k: v for k, v in row.items() if k not in drop_set} for row in rows]
    return cleaned, empty_cols


def clean_rows(rows: list[dict], col_types: dict[str, str]) -> tuple[list[dict], int, int]:
    """Clean rows in place. Returns (cleaned_rows, rows_dropped, duplicates_removed)."""
    if not rows:
        return [], 0, 0

    columns = list(rows[0].keys())
    rows_dropped = 0
    duplicates_removed = 0

    # Strip whitespace from all text values
    for row in rows:
        for col in columns:
            val = row.get(col)
            if isinstance(val, str):
                row[col] = val.strip()

    # Drop fully-empty rows
    cleaned: list[dict] = []
    for row in rows:
        if all(str(row.get(c, "")).strip() == "" for c in columns):
            rows_dropped += 1
        else:
            cleaned.append(row)

    # Drop exact duplicates (keep first occurrence)
    seen: set[tuple] = set()
    deduped: list[dict] = []
    for row in cleaned:
        key = tuple(str(row.get(c, "")) for c in columns)
        if key in seen:
            duplicates_removed += 1
        else:
            seen.add(key)
            deduped.append(row)

    # Coerce type mismatches → None
    pg_to_expected = {
        "BIGINT": "integer",
        "DOUBLE PRECISION": "float",
        "BOOLEAN": "boolean",
        "TIMESTAMP": "date",
    }
    for row in deduped:
        for col, pg_type in col_types.items():
            expected = pg_to_expected.get(pg_type)
            if expected and col in row:
                val = str(row[col]).strip()
                if val == "":
                    row[col] = None
                elif _try_parse_type(val) != expected:
                    detected = _try_parse_type(val)
                    # Allow integer values in float columns
                    if expected == "float" and detected == "integer":
                        continue
                    # Allow integer values (0/1) in boolean columns
                    if expected == "boolean" and detected == "integer" and val in ("0", "1"):
                        continue
                    row[col] = None

    # Drop rows where the primary key (first column) is NULL
    pk_col = columns[0]
    before = len(deduped)
    deduped = [
        row for row in deduped
        if row.get(pk_col) is not None and str(row.get(pk_col, "")).strip() != ""
    ]
    rows_dropped += before - len(deduped)

    return deduped, rows_dropped, duplicates_removed


# ---------------------------------------------------------------------------
# DB operations (raw SQL via SQLAlchemy text())
# ---------------------------------------------------------------------------

def _sanitize_table_name(name: str) -> str:
    """Sanitize a table name to prevent SQL injection."""
    return re.sub(r"[^a-zA-Z0-9_]", "_", name).lower().strip("_")[:63]


def _sanitize_col_name(name: str) -> str:
    """Sanitize a column name."""
    sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", name).lower().strip("_")
    return sanitized or "col"


def _has_unique_first_column(rows: list[dict]) -> bool:
    """Check if the first column has all unique non-empty values."""
    if not rows:
        return False
    first_col = list(rows[0].keys())[0]
    values = [str(row.get(first_col, "")).strip() for row in rows]
    non_empty = [v for v in values if v]
    return len(non_empty) == len(set(non_empty)) and len(non_empty) == len(rows)


async def create_table(
    session: AsyncSession, table_name: str, col_types: dict[str, str],
    use_serial_pk: bool = False,
) -> None:
    """CREATE TABLE IF NOT EXISTS with inferred PG types.

    If *use_serial_pk* is True, a synthetic ``_row_id SERIAL PRIMARY KEY``
    column is added and no data column is marked as PK.
    """
    safe_table = _sanitize_table_name(table_name)
    columns = list(col_types.items())
    if not columns:
        return

    col_defs = []
    if use_serial_pk:
        col_defs.append('"_row_id" SERIAL PRIMARY KEY')
        for col, pg_type in columns:
            col_defs.append(f'"{_sanitize_col_name(col)}" {pg_type}')
    else:
        for i, (col, pg_type) in enumerate(columns):
            safe_col = _sanitize_col_name(col)
            if i == 0:
                col_defs.append(f'"{safe_col}" {pg_type} PRIMARY KEY')
            else:
                col_defs.append(f'"{safe_col}" {pg_type}')

    ddl = f'CREATE TABLE IF NOT EXISTS "{safe_table}" ({", ".join(col_defs)})'
    await session.execute(text(ddl))
    await session.commit()


async def insert_rows(
    session: AsyncSession, table_name: str, rows: list[dict], col_types: dict[str, str],
    use_serial_pk: bool = False, batch_size: int = 200,
) -> int:
    """Bulk INSERT rows into the table. Returns count of rows inserted.

    When *use_serial_pk* is True the PK is auto-generated, so we use plain
    INSERT without ON CONFLICT (every row is unique by serial id).
    Rows are inserted in batches for performance.
    """
    if not rows:
        return 0

    safe_table = _sanitize_table_name(table_name)
    columns = list(col_types.keys())
    safe_cols = [_sanitize_col_name(c) for c in columns]

    col_list = ", ".join(f'"{c}"' for c in safe_cols)
    param_list = ", ".join(f":p{i}" for i in range(len(safe_cols)))

    if use_serial_pk:
        sql = f'INSERT INTO "{safe_table}" ({col_list}) VALUES ({param_list})'
    else:
        pk_col = safe_cols[0]
        if len(safe_cols) > 1:
            update_set = ", ".join(
                f'"{c}" = EXCLUDED."{c}"' for c in safe_cols[1:]
            )
            conflict = f'ON CONFLICT ("{pk_col}") DO UPDATE SET {update_set}'
        else:
            conflict = f'ON CONFLICT ("{pk_col}") DO NOTHING'
        sql = f'INSERT INTO "{safe_table}" ({col_list}) VALUES ({param_list}) {conflict}'

    inserted = 0
    # Batch inserts for performance
    for batch_start in range(0, len(rows), batch_size):
        batch = rows[batch_start:batch_start + batch_size]
        batch_params = []
        for row in batch:
            params = {}
            for i, col in enumerate(columns):
                val = row.get(col)
                if isinstance(val, str) and val.strip() == "":
                    val = None
                params[f"p{i}"] = val
            batch_params.append(params)
        await session.execute(text(sql), batch_params)
        inserted += len(batch)

    await session.commit()
    return inserted


# ---------------------------------------------------------------------------
# Gemini intelligence — auto metadata + context
# ---------------------------------------------------------------------------

METADATA_PROMPT = """\
You are a data engineer analyzing a newly uploaded dataset.

Table name: {table_name}
Columns: {columns}

Sample rows (first 5-10):
{sample_rows}

Generate a JSON response with:
1. "table_description": A clear 1-2 sentence description of what this table contains and its primary purpose.
2. "column_descriptions": An object mapping each column name to a 1-sentence description of what values it holds.
3. "business_context": A detailed 4-6 sentence paragraph covering ALL of the following:
   - What business domain or department this data serves (e.g., sales, finance, HR, marketing)
   - What KPIs or metrics can be calculated from these columns (name specific KPIs)
   - What analytical questions a business analyst could answer with this data (give 2-3 examples)
   - How this data could relate to or be joined with other common business data (e.g., "could be joined with customer data on customer_id")
   - Specific examples of actionable insights this data could reveal (not generic statements)

Return ONLY valid JSON, no markdown fences or explanation.
"""


async def generate_metadata_with_gemini(
    table_name: str, columns: list[str], rows: list[dict]
) -> dict[str, Any]:
    """Use Gemini to generate table description, column descriptions, and business context."""
    sample = rows[:10]
    sample_text = "\n".join(
        str({k: v for k, v in row.items()}) for row in sample
    )

    prompt = METADATA_PROMPT.format(
        table_name=table_name,
        columns=", ".join(columns),
        sample_rows=sample_text,
    )

    try:
        client = _get_client()
        response = client.models.generate_content(
            model=settings.gemini_model,
            contents=prompt,
        )
        raw = response.text.strip()
        # Strip markdown fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        return json.loads(raw)
    except Exception:
        logger.exception("Gemini metadata generation failed")
        return {
            "table_description": f"Uploaded table: {table_name}",
            "column_descriptions": {c: "" for c in columns},
            "business_context": f"Data uploaded as {table_name}.",
        }


# ---------------------------------------------------------------------------
# Main agent
# ---------------------------------------------------------------------------

class DataEngineerAgent:
    """Parses uploaded files, cleans data, loads to DB, and auto-generates metadata."""

    async def invoke(self, state: dict[str, Any]) -> dict[str, Any]:
        """Process an uploaded file end-to-end.

        Expected state keys:
            file_bytes: bytes — raw file content
            file_name: str — original filename
            db_session: AsyncSession
        """
        file_bytes: bytes = state["file_bytes"]
        file_name: str = state["file_name"]
        db_session: AsyncSession = state["db_session"]

        ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
        table_name = _sanitize_table_name(file_name.rsplit(".", 1)[0])

        # --- Parse ---
        if ext == "csv":
            rows = parse_csv(file_bytes)
        elif ext in ("xlsx", "xls"):
            rows = parse_excel(file_bytes)
        elif ext == "pdf":
            pdf_text = parse_pdf(file_bytes)
            # PDF is context-only — no tabular loading
            context_ids = await ingest_context(pdf_text, db_session, source=f"upload:{file_name}")
            gemini_meta = await generate_metadata_with_gemini(
                table_name, ["pdf_text"], [{"pdf_text": pdf_text[:500]}]
            )
            return {
                "table_name": table_name,
                "file_type": "pdf",
                "rows_total": 0,
                "rows_inserted": 0,
                "rows_dropped": 0,
                "duplicates_removed": 0,
                "issues": [],
                "metadata": {
                    "description": gemini_meta.get("table_description", ""),
                    "columns": [],
                },
                "context_generated": gemini_meta.get("business_context", ""),
                "context_ids": context_ids,
            }
        else:
            raise ValueError(f"Unsupported file type: .{ext}")

        if not rows:
            return {
                "table_name": table_name,
                "file_type": ext,
                "rows_total": 0,
                "rows_inserted": 0,
                "rows_dropped": 0,
                "duplicates_removed": 0,
                "issues": [{"column": "*", "issue": "File is empty", "action": "none"}],
                "metadata": {"description": "", "columns": []},
                "context_generated": "",
            }

        rows_total = len(rows)
        issues: list[dict] = []

        # --- Drop entirely-empty columns ---
        rows, dropped_cols = _drop_empty_columns(rows)
        for col_name in dropped_cols:
            issues.append({
                "column": col_name,
                "issue": "column is entirely empty",
                "action": "dropped_column",
            })

        # --- Type inference ---
        col_types = infer_column_types(rows)

        # --- Hygiene checks ---
        issues.extend(check_hygiene(rows))

        # --- Clean ---
        cleaned, rows_dropped, duplicates_removed = clean_rows(rows, col_types)

        # --- DB load ---
        # Use synthetic serial PK if first column has duplicate values
        use_serial_pk = not _has_unique_first_column(cleaned)
        if use_serial_pk:
            logger.info("First column is not unique — using serial _row_id PK for %s", table_name)
        await create_table(db_session, table_name, col_types, use_serial_pk=use_serial_pk)
        rows_inserted = await insert_rows(db_session, table_name, cleaned, col_types, use_serial_pk=use_serial_pk)

        # --- Gemini metadata ---
        columns = list(col_types.keys())
        gemini_meta = await generate_metadata_with_gemini(table_name, columns, cleaned)

        # Build column metadata for metadata_store
        col_descriptions = gemini_meta.get("column_descriptions", {})
        columns_metadata = []
        if use_serial_pk:
            columns_metadata.append({
                "name": "_row_id",
                "type": "SERIAL",
                "description": "Auto-generated row identifier (primary key)",
            })
        for col, pg_type in col_types.items():
            columns_metadata.append({
                "name": _sanitize_col_name(col),
                "type": pg_type,
                "description": col_descriptions.get(col, ""),
            })

        table_description = gemini_meta.get("table_description", f"Uploaded: {table_name}")
        business_context = gemini_meta.get("business_context", "")

        # Store metadata
        await metadata_store.upsert(
            db_session,
            table_name=table_name,
            description=table_description,
            columns_metadata=columns_metadata,
        )

        # Store business context
        if business_context:
            await ingest_context(business_context, db_session, source=f"upload:{file_name}")

        return {
            "table_name": table_name,
            "file_type": ext,
            "rows_total": rows_total,
            "rows_inserted": rows_inserted,
            "rows_dropped": rows_dropped,
            "duplicates_removed": duplicates_removed,
            "issues": issues,
            "metadata": {
                "description": table_description,
                "columns": columns_metadata,
            },
            "context_generated": business_context,
        }
